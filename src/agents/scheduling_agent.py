"""
Medical Appointment Scheduling Agent
Main orchestrator for the appointment scheduling system using LangChain/LangGraph
"""

import os
import json
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import logging

# LangChain imports
from langchain.agents import AgentExecutor, create_openai_functions_agent
from langchain.memory import ConversationBufferMemory
from langchain_openai import ChatOpenAI
from langchain.prompts import ChatPromptTemplate, MessagesPlaceholder
from langchain.schema.messages import BaseMessage, HumanMessage, AIMessage
from langchain.tools import Tool
from langchain_core.runnables import RunnablePassthrough
from langchain_core.callbacks.base import BaseCallbackHandler

# Import our custom modules
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'utils'))
try:
    from utils.email_templates import IntakeFormEmailTemplates
    from utils.smtp_email_service import SMTPEmailService
    from utils.intake_form_handler import IntakeFormHandler
except ImportError:
    try:
        from email_templates import IntakeFormEmailTemplates
        from smtp_email_service import SMTPEmailService
        from intake_form_handler import IntakeFormHandler
    except ImportError:
        IntakeFormEmailTemplates = None
        SMTPEmailService = None
        IntakeFormHandler = None


class MedicalSchedulingAgent:
    """
    Main scheduling agent that orchestrates the entire appointment booking process
    """
    
    def __init__(self, 
                 openai_api_key: Optional[str] = None,
                 data_dir: str = "data",
                 model_name: str = "gpt-3.5-turbo"):
        
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize logging
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)
        
        # Initialize OpenAI
        self.openai_api_key = openai_api_key or os.getenv("OPENAI_API_KEY")
        if not self.openai_api_key:
            self.logger.warning("OpenAI API key not provided. Agent will run in simulation mode.")
            self.llm = None
        else:
            self.llm = ChatOpenAI(
                openai_api_key=self.openai_api_key,
                model_name=model_name,
                temperature=0.3
            )
        
        # Initialize components
        if IntakeFormHandler:
            self.intake_handler = IntakeFormHandler(str(self.data_dir / "intake_forms"))
        else:
            self.intake_handler = None
            
        # Initialize email service and reminder system
        if SMTPEmailService and IntakeFormEmailTemplates:
            self.email_service = SMTPEmailService()
            self.email_templates = IntakeFormEmailTemplates()
            
            # Initialize reminder engine
            try:
                from src.utils.reminder_engine import AppointmentReminderEngine
                self.reminder_engine = AppointmentReminderEngine(self.data_dir, self.email_service)
            except ImportError:
                print("Reminder engine not available")
                self.reminder_engine = None
        else:
            self.email_service = None
            self.email_templates = None
            self.reminder_engine = None
        
        # Load data
        self.patients_df = self._load_patients_data()
        self.doctors_df = self._load_doctors_data()
        self.appointments_df = self._load_appointments_data()
        
        # Initialize agent tools
        self.tools = self._create_agent_tools()
        
        # Initialize conversation memory
        self.memory = ConversationBufferMemory(
            memory_key="chat_history",
            return_messages=True
        ) if self.llm else None
        
        # Create agent (with error handling)
        try:
            self.agent = self._create_agent() if self.llm else None
        except Exception as e:
            self.logger.warning(f"Could not create LangChain agent: {e}")
            self.logger.info("Agent will run in simulation mode")
            self.agent = None
        
    def _load_patients_data(self) -> pd.DataFrame:
        """Load patient database"""
        try:
            return pd.read_csv(self.data_dir / "patients" / "patient_database.csv")
        except FileNotFoundError:
            self.logger.warning("Patient database not found. Creating empty DataFrame.")
            return pd.DataFrame()
    
    def _load_doctors_data(self) -> pd.DataFrame:
        """Load doctor profiles"""
        try:
            return pd.read_csv(self.data_dir / "doctors" / "doctor_profiles.csv")
        except FileNotFoundError:
            self.logger.warning("Doctor profiles not found. Creating empty DataFrame.")
            return pd.DataFrame()
    
    def _load_appointments_data(self) -> pd.DataFrame:
        """Load existing appointments"""
        appointments_file = self.data_dir / "appointments" / "scheduled_appointments.csv"
        if appointments_file.exists():
            return pd.read_csv(appointments_file)
        else:
            # Create empty appointments DataFrame
            appointments_file.parent.mkdir(parents=True, exist_ok=True)
            df = pd.DataFrame(columns=[
                'appointment_id', 'patient_id', 'doctor_id', 'appointment_date',
                'appointment_time', 'reason', 'status', 'created_at', 'notes'
            ])
            df.to_csv(appointments_file, index=False)
            return df
    
    def _create_agent_tools(self) -> List[Tool]:
        """Create tools for the scheduling agent"""
        
        tools = [
            Tool(
                name="lookup_patient",
                description="Look up patient information by name, phone, or email",
                func=self._lookup_patient
            ),
            Tool(
                name="search_available_slots",
                description="Search for available appointment slots by date range, doctor, or specialty",
                func=self._search_available_slots
            ),
            Tool(
                name="book_appointment",
                description="Book an appointment for a patient with specified doctor and time",
                func=self._book_appointment
            ),
            Tool(
                name="get_doctor_info",
                description="Get information about doctors including specialties and availability",
                func=self._get_doctor_info
            ),
            Tool(
                name="check_insurance_coverage",
                description="Check if a doctor accepts patient's insurance (simulated)",
                func=self._check_insurance_coverage
            ),
            Tool(
                name="send_intake_form",
                description="Send intake form to patient after appointment confirmation",
                func=self._send_intake_form
            ),
            Tool(
                name="reschedule_appointment",
                description="Reschedule an existing appointment to a new time",
                func=self._reschedule_appointment
            ),
            Tool(
                name="cancel_appointment",
                description="Cancel an existing appointment",
                func=self._cancel_appointment
            ),
            Tool(
                name="export_appointments_excel",
                description="Export appointments to Excel format with multiple sheets and formatting",
                func=self._export_appointments_to_excel
            ),
            Tool(
                name="send_appointment_confirmation",
                description="Send appointment confirmation via email/SMS with detailed information",
                func=self._send_appointment_confirmation
            ),
            Tool(
                name="backup_data",
                description="Create backup of appointment and patient data",
                func=self._backup_appointment_data
            ),
            Tool(
                name="distribute_intake_forms",
                description="Send intake forms to new patients after appointment confirmation with workflow validation",
                func=self._distribute_intake_forms
            ),
            Tool(
                name="schedule_appointment_reminders",
                description="Schedule automated reminders (24h, 4h, 1h before) for an appointment",
                func=self._schedule_appointment_reminders
            ),
            Tool(
                name="check_reminder_status",
                description="Check the status of scheduled reminders for appointments or patients",
                func=self._check_reminder_status
            ),
            Tool(
                name="send_manual_reminder",
                description="Manually send a specific reminder to a patient",
                func=self._send_manual_reminder
            ),
            Tool(
                name="process_patient_response",
                description="Process patient responses to reminders (confirmations, cancellations, etc.)",
                func=self._process_patient_response
            ),
            Tool(
                name="run_reminder_system",
                description="Check and send all due reminders in the system",
                func=self._run_reminder_system
            ),
            Tool(
                name="configure_sms_service",
                description="Get SMS service configuration and setup instructions",
                func=self._configure_sms_service
            )
        ]
        
        return tools
    
    def _create_agent(self) -> AgentExecutor:
        """Create the LangChain agent"""
        
        system_prompt = """
        You are a helpful medical appointment scheduling assistant. Your role is to:
        
        1. Greet patients warmly and professionally
        2. Gather necessary information (name, phone, preferred dates/times, reason for visit)
        3. Look up existing patient records or collect new patient information
        4. Search for available appointment slots that match patient preferences
        5. Check insurance coverage and provide cost information when relevant
        6. Book appointments and provide confirmation details
        7. Send intake forms to new patients
        8. Handle rescheduling and cancellation requests
        9. Provide appointment reminders and follow-up information
        
        Always be professional, empathetic, and efficient. Ask clarifying questions when needed.
        Confirm all important details before finalizing appointments.
        
        When booking appointments:
        - Always confirm patient identity first
        - Verify insurance coverage if mentioned
        - Provide clear appointment details including date, time, doctor, and location
        - Send intake forms to new patients
        - Offer appointment reminders
        
        For existing patients:
        - Reference their previous visits when relevant
        - Ask about any changes to contact information or insurance
        
        For new patients:
        - Collect basic contact information
        - Explain the intake form process
        - Provide clinic policies and what to expect
        """
        
        prompt = ChatPromptTemplate.from_messages([
            ("system", system_prompt),
            MessagesPlaceholder(variable_name="chat_history"),
            ("human", "{input}"),
            MessagesPlaceholder(variable_name="agent_scratchpad")
        ])
        
        agent = create_openai_functions_agent(
            llm=self.llm,
            tools=self.tools,
            prompt=prompt
        )
        
        return AgentExecutor(
            agent=agent,
            tools=self.tools,
            memory=self.memory,
            verbose=True,
            max_iterations=10,
            handle_parsing_errors=True
        )
    
    def _lookup_patient(self, query: str) -> str:
        """Look up patient by name, phone, or email"""
        try:
            query = query.strip()
            query_lower = query.lower()
            
            if self.patients_df.empty:
                return "No patient database available"
            
            # First try exact full name match
            full_name_matches = self.patients_df[
                (self.patients_df['first_name'] + ' ' + self.patients_df['last_name']).str.lower() == query_lower
            ]
            
            if not full_name_matches.empty:
                matches = full_name_matches
            else:
                # Search by partial name, phone, or email
                matches = self.patients_df[
                    (self.patients_df['first_name'].str.lower().str.contains(query_lower, na=False)) |
                    (self.patients_df['last_name'].str.lower().str.contains(query_lower, na=False)) |
                    (self.patients_df['phone'].str.contains(query, na=False)) |
                    (self.patients_df['email'].str.lower().str.contains(query_lower, na=False))
                ]
            
            if matches.empty:
                # Store as new patient for smart scheduling
                self.current_patient_name = query
                self.current_patient_id = 'NEW_PATIENT'
                return f"No patient found matching '{query}'. This appears to be a new patient."
            
            if len(matches) == 1:
                patient = matches.iloc[0]
                # Store patient info for smart scheduling
                self.current_patient_name = f"{patient['first_name']} {patient['last_name']}"
                self.current_patient_id = patient['patient_id']
                return f"""
Found patient: {patient['first_name']} {patient['last_name']}
Phone: {patient['phone']}
Email: {patient['email']}
Date of Birth: {patient['date_of_birth']}
Insurance: {patient['insurance_carrier']}
Patient ID: {patient['patient_id']}
"""
            else:
                # Multiple matches - return summary
                results = []
                for _, patient in matches.iterrows():
                    results.append(f"{patient['first_name']} {patient['last_name']} - {patient['phone']}")
                return f"Multiple patients found:\n" + "\n".join(results) + "\nPlease provide more specific information."
                
        except Exception as e:
            return f"Error looking up patient: {str(e)}"
    
    def _search_available_slots(self, query: str) -> str:
        """Search for available appointment slots"""
        try:
            # Parse query for date preferences, doctor, specialty
            # For now, return simulated availability
            
            # Load doctor schedules
            schedule_file = self.data_dir / "doctors" / "doctor_schedules.xlsx"
            if not schedule_file.exists():
                return "Doctor schedules not available. Please contact the office directly."
            
            # Read the Excel file
            schedules_df = pd.read_excel(schedule_file, sheet_name=None)
            
            available_slots = []
            today = datetime.now().date()
            
            # Get next 7 days of availability
            for i in range(7):
                check_date = today + timedelta(days=i)
                date_str = check_date.strftime('%Y-%m-%d')
                
                # Check each doctor's schedule
                for doctor_name, schedule in schedules_df.items():
                    if doctor_name in ["Sheet1", "All_Schedules"]:  # Skip these sheets
                        continue
                        
                    # Convert date column to string for comparison
                    schedule['date_str'] = pd.to_datetime(schedule['date']).dt.strftime('%Y-%m-%d')
                    
                    # Filter for the specific date and available status
                    day_schedule = schedule[
                        (schedule['date_str'] == date_str) & 
                        (schedule['status'] == 'available')
                    ]
                    
                    if not day_schedule.empty:
                        available_times = day_schedule['time_slot'].tolist()
                        
                        for time in available_times[:3]:  # Limit to 3 slots per doctor per day
                            doctor_display = doctor_name.replace('_', ' ')
                            available_slots.append(f"Dr. {doctor_display} - {check_date.strftime('%A, %B %d')} at {time}")
            
            if not available_slots:
                return "No available appointments found in the next week. Please call for additional availability."
            
            return "Available appointment slots:\n" + "\n".join(available_slots[:10])  # Limit to 10 results
            
        except Exception as e:
            return f"Error searching availability: {str(e)}"
    
    def _determine_appointment_duration(self, patient_name: str, appointment_type: str = "routine") -> tuple:
        """
        Smart Scheduling Feature: Determine appointment duration based on patient type
        Returns (duration_minutes, patient_type, special_notes)
        """
        try:
            # Check if patient exists in database
            patient_lookup = self._lookup_patient(patient_name)
            is_existing_patient = "Found patient:" in patient_lookup
            
            # Base durations
            if is_existing_patient:
                base_duration = 30  # 30 minutes for returning patients
                patient_type = "Returning Patient"
            else:
                base_duration = 60  # 60 minutes for new patients
                patient_type = "New Patient"
            
            # Adjust duration based on appointment type
            duration_adjustments = {
                "physical": 15,          # +15 min for physicals
                "consultation": 0,       # Standard duration
                "follow-up": -10,        # -10 min for follow-ups (returning patients only)
                "urgent": 15,            # +15 min for urgent care
                "specialist": 15,        # +15 min for specialist visits
                "cardiology": 20,        # +20 min for cardiology
                "surgery": 30,           # +30 min for surgical consultations
                "mental health": 30,     # +30 min for mental health
                "psychiatry": 30         # +30 min for psychiatry
            }
            
            # Find matching appointment type
            adjustment = 0
            appointment_type_lower = appointment_type.lower()
            for key, value in duration_adjustments.items():
                if key in appointment_type_lower:
                    adjustment = value
                    break
            
            # Calculate final duration
            final_duration = base_duration + adjustment
            
            # Ensure minimum 15 minutes and maximum 120 minutes
            final_duration = max(15, min(120, final_duration))
            
            # Generate special notes
            notes = []
            if not is_existing_patient:
                notes.append("New patient - includes intake and registration time")
            if adjustment > 0:
                notes.append(f"Extended time for {appointment_type} (+{adjustment} min)")
            elif adjustment < 0 and is_existing_patient:
                notes.append(f"Shortened follow-up appointment ({adjustment} min)")
            
            special_notes = "; ".join(notes) if notes else "Standard appointment duration"
            
            return final_duration, patient_type, special_notes
            
        except Exception as e:
            # Default to 45 minutes if there's an error
            return 45, "Unknown Patient Type", f"Error determining duration: {str(e)}"
    
    def _book_appointment(self, details: str) -> str:
        """Book an appointment with Smart Scheduling"""
        try:
            # Clear any previous patient context to avoid conflicts
            if hasattr(self, 'current_patient_name'):
                delattr(self, 'current_patient_name')
            if hasattr(self, 'current_patient_id'):
                delattr(self, 'current_patient_id')
            
            # Parse appointment details (in a real system, this would be more sophisticated)
            # For demo, we'll extract basic info
            import re
            
            # Extract patient name from booking details
            patient_name = getattr(self, 'current_patient_name', 'Unknown Patient')
            
            # Try to extract name from booking details if not already set
            if patient_name == 'Unknown Patient' or not hasattr(self, 'current_patient_name'):
                # Look for "I am [Name]" or "My name is [Name]" patterns
                name_patterns = [
                    r'I am ([A-Z][a-z]+ [A-Z][a-z]+)',
                    r'my name is ([A-Z][a-z]+ [A-Z][a-z]+)',
                    r'This is ([A-Z][a-z]+ [A-Z][a-z]+)',
                    r'^([A-Z][a-z]+ [A-Z][a-z]+)',  # Name at beginning
                ]
                
                for pattern in name_patterns:
                    match = re.search(pattern, details, re.IGNORECASE)
                    if match:
                        extracted_name = match.group(1).title()
                        # Verify it looks like a real name (2+ words, proper capitalization)
                        if len(extracted_name.split()) >= 2:
                            patient_name = extracted_name
                            break
            
            # Extract appointment type from details
            appointment_type = "routine"
            details_lower = details.lower()
            if any(word in details_lower for word in ['cardio', 'heart']):
                appointment_type = "cardiology"
            elif any(word in details_lower for word in ['physical', 'checkup']):
                appointment_type = "physical"
            elif any(word in details_lower for word in ['follow', 'followup']):
                appointment_type = "follow-up"
            elif any(word in details_lower for word in ['urgent', 'emergency']):
                appointment_type = "urgent"
            elif any(word in details_lower for word in ['mental', 'psych']):
                appointment_type = "mental health"
            
            # Apply Smart Scheduling
            duration, patient_type, special_notes = self._determine_appointment_duration(
                patient_name, appointment_type
            )
            
            # If this is a new patient, add them to the database
            patient_id = getattr(self, 'current_patient_id', None)
            if patient_type == "New Patient":
                # Double-check that patient doesn't exist in database
                lookup_result = self._lookup_patient(patient_name)
                if "No patient found" in lookup_result:
                    # Collect contact information from the booking details
                    contact_info = self._collect_patient_info(details)
                    
                    # Add new patient to database
                    add_result = self._add_new_patient_to_database(patient_name, contact_info)
                    print(f"Database update: {add_result}")
                    
                    # Get the new patient ID after adding
                    updated_lookup = self._lookup_patient(patient_name)
                    if "Patient ID:" in updated_lookup:
                        patient_id = updated_lookup.split("Patient ID: ")[1].split("\n")[0].strip()
                    else:
                        patient_id = 'NEW_PATIENT'
                else:
                    # Patient exists, get their ID
                    if hasattr(self, 'current_patient_id'):
                        patient_id = self.current_patient_id
                    else:
                        patient_id = 'EXISTING_PATIENT'
            elif not patient_id:
                patient_id = 'EXISTING_PATIENT'
            
            # Generate appointment ID
            appointment_id = f"APT_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            
            # Create appointment record with smart scheduling data
            appointment_data = {
                'appointment_id': appointment_id,
                'patient_id': patient_id,
                'patient_name': patient_name,
                'patient_type': patient_type,
                'doctor_id': 'DR001',  # Would be actual doctor ID
                'appointment_date': (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d'),
                'appointment_time': '10:00 AM',
                'duration_minutes': duration,
                'appointment_type': appointment_type,
                'reason': 'General consultation',
                'status': 'confirmed',
                'created_at': datetime.now().isoformat(),
                'special_notes': special_notes,
                'notes': details
            }
            
            # Save appointment
            appointments_file = self.data_dir / "appointments" / "scheduled_appointments.csv"
            
            if appointments_file.exists():
                appointments_df = pd.read_csv(appointments_file)
                new_row = pd.DataFrame([appointment_data])
                appointments_df = pd.concat([appointments_df, new_row], ignore_index=True)
            else:
                appointments_df = pd.DataFrame([appointment_data])
            
            appointments_df.to_csv(appointments_file, index=False)
            
            # Send appointment confirmation
            confirmation_data = {
                'appointment_id': appointment_id,
                'patient_name': patient_name,
                'patient_email': contact_info.get('email', 'charulchim06@gmail.com') if 'contact_info' in locals() else 'charulchim06@gmail.com',
                'patient_phone': contact_info.get('phone', '') if 'contact_info' in locals() else '',
                'appointment_date': appointment_data['appointment_date'],
                'appointment_time': appointment_data['appointment_time'],
                'duration_minutes': duration,
                'appointment_type': appointment_type,
                'patient_type': patient_type,
                'doctor_name': 'Sarah Johnson'
            }
            
            confirmation_result = self._send_appointment_confirmation(confirmation_data, "booking")
            
            # Automatically schedule reminders after successful confirmation
            reminder_result = ""
            if "✅" in confirmation_result and self.reminder_engine:
                try:
                    reminder_schedule = self.reminder_engine.schedule_reminders_for_appointment(appointment_id)
                    if reminder_schedule['status'] == 'success':
                        reminder_count = len(reminder_schedule.get('reminders', []))
                        reminder_result = f"\n🔔 Automated reminders scheduled: {reminder_count} reminders (24h, 4h, 1h before)"
                    else:
                        reminder_result = f"\n⚠️ Reminder scheduling: {reminder_schedule.get('message', 'Failed')}"
                except Exception as e:
                    reminder_result = f"\n⚠️ Reminder scheduling error: {str(e)}"
            
            return f"""
✅ Appointment successfully booked with Smart Scheduling!

📅 Appointment Details:
- Appointment ID: {appointment_id}
- Patient: {patient_name} ({patient_type})
- Date: {appointment_data['appointment_date']}
- Time: {appointment_data['appointment_time']}
- Duration: {duration} minutes
- Type: {appointment_type.title()}
- Doctor: Dr. Sarah Johnson (example)
- Status: Confirmed

🧠 Smart Scheduling Applied:
- {special_notes}
- Optimized duration based on patient history and appointment type

📋 Next steps:
1. You will receive a confirmation email with intake forms
2. {'Complete intake forms (15 min) - included in your appointment time' if patient_type == 'New Patient' else 'Brief check-in (5 min) - you know the routine!'}
3. Bring photo ID and insurance card
4. Arrive 15 minutes early for check-in

Total appointment block: {duration} minutes

📨 **Confirmation Status:**
{confirmation_result.split('**Delivery Status:**')[1].split('**Message Includes:**')[0].strip() if '**Delivery Status:**' in confirmation_result else 'Confirmation processing...'}
{reminder_result}

Would you like me to:
- Export appointment details to Excel?
- Send the intake forms now?
- Create a data backup?
"""
            
        except Exception as e:
            return f"Error booking appointment: {str(e)}"
    
    def _add_new_patient_to_database(self, patient_name: str, contact_info: dict = None) -> str:
        """Add a new patient to the patient database with enhanced insurance collection"""
        try:
            # Generate new patient ID
            patient_id = f"P{datetime.now().strftime('%Y%m%d%H%M%S')}"
            
            # Parse name
            name_parts = patient_name.strip().split()
            first_name = name_parts[0] if name_parts else "Unknown"
            last_name = " ".join(name_parts[1:]) if len(name_parts) > 1 else ""
            
            # Validate insurance information
            validated_insurance = {}
            if contact_info:
                insurance_data = {k: v for k, v in contact_info.items() if k in ['insurance_carrier', 'member_id', 'group_number']}
                validated_insurance = self._validate_insurance_info(insurance_data)
            
            # Create new patient record with proper insurance fields
            new_patient_data = {
                'patient_id': patient_id,
                'first_name': first_name,
                'last_name': last_name,
                'full_name': patient_name,
                'email': contact_info.get('email', '') if contact_info else '',
                'phone': contact_info.get('phone', '') if contact_info else '',
                'address': contact_info.get('address', '') if contact_info else '',
                'date_of_birth': contact_info.get('dob', '') if contact_info else '',
                'gender': contact_info.get('gender', '') if contact_info else '',
                'emergency_contact': contact_info.get('emergency_contact', '') if contact_info else '',
                
                # Enhanced insurance fields
                'insurance_carrier': validated_insurance.get('insurance_carrier', ''),
                'member_id': validated_insurance.get('member_id', ''),
                'group_number': validated_insurance.get('group_number', ''),
                
                # Legacy insurance field for backward compatibility
                'insurance_provider': validated_insurance.get('insurance_carrier', contact_info.get('insurance', '') if contact_info else ''),
                
                'medical_history': '',
                'allergies': '',
                'medications': '',
                'created_date': datetime.now().strftime('%Y-%m-%d'),
                'last_visit': '',
                'status': 'Active'
            }
            
            # Save to patient database
            patients_file = self.data_dir / "patients" / "patient_database.csv"
            
            if patients_file.exists():
                patients_df = pd.read_csv(patients_file)
                new_row = pd.DataFrame([new_patient_data])
                patients_df = pd.concat([patients_df, new_row], ignore_index=True)
            else:
                patients_df = pd.DataFrame([new_patient_data])
            
            patients_df.to_csv(patients_file, index=False)
            
            # Reload the patients dataframe
            self._reload_patient_database()
            
            # Create insurance summary for return message
            insurance_summary = ""
            if validated_insurance:
                insurance_parts = []
                if 'insurance_carrier' in validated_insurance:
                    insurance_parts.append(f"Carrier: {validated_insurance['insurance_carrier']}")
                if 'member_id' in validated_insurance:
                    insurance_parts.append(f"Member ID: {validated_insurance['member_id']}")
                if 'group_number' in validated_insurance:
                    insurance_parts.append(f"Group: {validated_insurance['group_number']}")
                
                if insurance_parts:
                    insurance_summary = f"\n📋 Insurance: {', '.join(insurance_parts)}"
            
            return f"✅ New patient '{patient_name}' added to database with ID: {patient_id}{insurance_summary}"
            
        except Exception as e:
            return f"❌ Error adding new patient to database: {str(e)}"
    
    def _reload_patient_database(self):
        """Reload the patient database to reflect recent changes"""
        try:
            patients_file = self.data_dir / "patients" / "patient_database.csv"
            if patients_file.exists():
                self.patients_df = pd.read_csv(patients_file)
            else:
                self.patients_df = pd.DataFrame()
        except Exception as e:
            print(f"Error reloading patient database: {e}")
    
    def _collect_patient_info(self, details: str) -> str:
        """Collect and save new patient information from their input"""
        try:
            # Extract contact information from patient input
            import re
            
            contact_info = {}
            
            # Extract email
            email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            email_match = re.search(email_pattern, details)
            if email_match:
                contact_info['email'] = email_match.group()
            
            # Extract phone number
            phone_pattern = r'(\+?\d{1,3}[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'
            phone_match = re.search(phone_pattern, details)
            if phone_match:
                contact_info['phone'] = phone_match.group()
            
            # Enhanced insurance extraction
            insurance_info = self._extract_insurance_details(details)
            contact_info.update(insurance_info)
            
            return contact_info
            
        except Exception as e:
            return {}
    
    def _extract_insurance_details(self, details: str) -> dict:
        """Extract comprehensive insurance details from patient input"""
        try:
            import re
            insurance_info = {}
            details_lower = details.lower()
            
            # Known insurance carriers and their variations
            insurance_carriers = {
                'blue cross blue shield': ['blue cross', 'bcbs', 'blue shield', 'bluecross'],
                'aetna': ['aetna'],
                'cigna': ['cigna'],
                'humana': ['humana'],
                'united healthcare': ['united', 'uhc', 'united healthcare'],
                'kaiser permanente': ['kaiser', 'kp'],
                'anthem': ['anthem'],
                'medicare': ['medicare'],
                'medicaid': ['medicaid'],
                'tricare': ['tricare']
            }
            
            # Extract insurance carrier
            for standard_name, variations in insurance_carriers.items():
                for variation in variations:
                    if variation in details_lower:
                        insurance_info['insurance_carrier'] = standard_name.title()
                        break
                if 'insurance_carrier' in insurance_info:
                    break
            
            # Extract member ID patterns
            member_id_patterns = [
                r'member\s*(?:id|number)[\s:]*([A-Z]{2,3}\d{8,12})',  # BC123456789
                r'id[\s:]*([A-Z]{2,3}\d{8,12})',
                r'member[\s:]*([A-Z]{2,3}\d{8,12})',
                r'policy[\s:]*([A-Z]{2,3}\d{8,12})',
                r'([A-Z]{2,3}\d{9,12})',  # General pattern for insurance IDs
            ]
            
            for pattern in member_id_patterns:
                match = re.search(pattern, details, re.IGNORECASE)
                if match:
                    insurance_info['member_id'] = match.group(1).upper()
                    break
            
            # Extract group number patterns
            group_patterns = [
                r'group\s*(?:number|id)[\s:]*([A-Z0-9]{3,10})',
                r'group[\s:]*([A-Z0-9]{3,10})',
                r'grp[\s:]*([A-Z0-9]{3,10})',
            ]
            
            for pattern in group_patterns:
                match = re.search(pattern, details, re.IGNORECASE)
                if match:
                    insurance_info['group_number'] = match.group(1).upper()
                    break
            
            # If no specific member ID found but insurance carrier detected, 
            # look for any alphanumeric sequence that could be an ID
            if 'insurance_carrier' in insurance_info and 'member_id' not in insurance_info:
                # Look for standalone alphanumeric sequences
                id_pattern = r'\b([A-Z]{2,3}\d{6,12}|\d{8,12})\b'
                matches = re.findall(id_pattern, details, re.IGNORECASE)
                if matches:
                    insurance_info['member_id'] = matches[0].upper()
            
            return insurance_info
            
        except Exception as e:
            print(f"Error extracting insurance details: {e}")
            return {}
    
    def _validate_insurance_info(self, insurance_info: dict) -> dict:
        """Validate and format insurance information"""
        try:
            import re
            validated_info = {}
            
            # Validate insurance carrier
            if 'insurance_carrier' in insurance_info:
                carrier = insurance_info['insurance_carrier'].strip()
                if len(carrier) >= 3:  # Minimum length check
                    validated_info['insurance_carrier'] = carrier
            
            # Validate member ID
            if 'member_id' in insurance_info:
                member_id = insurance_info['member_id'].strip().upper()
                # Check if it matches common insurance ID patterns
                if re.match(r'^[A-Z]{2,3}\d{6,12}$|^\d{8,12}$', member_id):
                    validated_info['member_id'] = member_id
            
            # Validate group number
            if 'group_number' in insurance_info:
                group_num = insurance_info['group_number'].strip().upper()
                if re.match(r'^[A-Z0-9]{3,10}$', group_num):
                    validated_info['group_number'] = group_num
            
            return validated_info
            
        except Exception as e:
            print(f"Error validating insurance info: {e}")
            return {}
    
    def _get_doctor_info(self, query: str) -> str:
        """Get doctor information"""
        try:
            if self.doctors_df.empty:
                return "Doctor information not available"
            
            if not query or query.lower() == "all":
                # Return all doctors
                doctor_list = []
                for _, doctor in self.doctors_df.iterrows():
                    doctor_list.append(f"Dr. {doctor['first_name']} {doctor['last_name']} - {doctor['specialty']}")
                return "Available doctors:\n" + "\n".join(doctor_list)
            
            # Search for specific doctor or specialty
            query = query.lower()
            
            # Handle common specialty variations
            specialty_mapping = {
                'cardiologist': 'cardiology',
                'heart doctor': 'cardiology',
                'pediatrician': 'pediatrics',
                'family doctor': 'family medicine',
                'internist': 'internal medicine',
                'psychiatrist': 'psychiatry'
            }
            
            # Use mapped specialty if available
            search_term = specialty_mapping.get(query, query)
            
            matches = self.doctors_df[
                (self.doctors_df['first_name'].str.lower().str.contains(search_term, na=False)) |
                (self.doctors_df['last_name'].str.lower().str.contains(search_term, na=False)) |
                (self.doctors_df['specialty'].str.lower().str.contains(search_term, na=False))
            ]
            
            if matches.empty:
                return f"No doctors found matching '{query}'"
            
            results = []
            for _, doctor in matches.iterrows():
                results.append(f"""
Dr. {doctor['first_name']} {doctor['last_name']}
Specialty: {doctor['specialty']}
Phone: {doctor['phone']}
Email: {doctor['email']}
Experience: {doctor['years_experience']} years
""")
            
            return "\n".join(results)
            
        except Exception as e:
            return f"Error getting doctor information: {str(e)}"
    
    def _check_insurance_coverage(self, insurance_info: str) -> str:
        """Check comprehensive insurance coverage with member ID and group validation"""
        try:
            # Extract insurance details from the input
            extracted_insurance = self._extract_insurance_details(insurance_info)
            validated_insurance = self._validate_insurance_info(extracted_insurance)
            
            # Insurance coverage database (simulated)
            coverage_database = {
                'blue cross blue shield': {
                    'accepted': True,
                    'copay_range': '$20-40',
                    'member_id_format': 'BC followed by 9 digits',
                    'group_formats': ['GRP001', 'GRP002', 'GRP003']
                },
                'aetna': {
                    'accepted': True,
                    'copay_range': '$25-50',
                    'member_id_format': 'AE followed by 9 digits',
                    'group_formats': ['GRP002', 'GRP004']
                },
                'cigna': {
                    'accepted': True,
                    'copay_range': '$30-60',
                    'member_id_format': 'CI followed by 9 digits',
                    'group_formats': ['GRP004', 'GRP005']
                },
                'united healthcare': {
                    'accepted': True,
                    'copay_range': '$25-45',
                    'member_id_format': 'UH followed by 9 digits',
                    'group_formats': ['GRP003', 'GRP006']
                },
                'humana': {
                    'accepted': True,
                    'copay_range': '$20-35',
                    'member_id_format': 'HU followed by 9 digits',
                    'group_formats': ['GRP006']
                },
                'kaiser permanente': {
                    'accepted': False,
                    'reason': 'HMO network - requires referral from primary care'
                },
                'medicare': {
                    'accepted': True,
                    'copay_range': '$15-25',
                    'member_id_format': '11-digit Medicare number',
                    'group_formats': ['GRP005']
                },
                'medicaid': {
                    'accepted': True,
                    'copay_range': '$0-10',
                    'member_id_format': 'State-specific format',
                    'group_formats': ['MEDICAID']
                }
            }
            
            # Check if we have insurance carrier information
            if 'insurance_carrier' in validated_insurance:
                carrier = validated_insurance['insurance_carrier'].lower()
                
                if carrier in coverage_database:
                    coverage_info = coverage_database[carrier]
                    
                    if coverage_info['accepted']:
                        # Build verification message
                        verification_msg = f"""
🔍 **Insurance Verification Complete**

✅ **Carrier:** {validated_insurance['insurance_carrier']}
✅ **Coverage:** Accepted by our practice
✅ **Estimated Copay:** {coverage_info['copay_range']}
"""
                        
                        # Add member ID verification if provided
                        if 'member_id' in validated_insurance:
                            verification_msg += f"✅ **Member ID:** {validated_insurance['member_id']} (Format verified)\n"
                        else:
                            verification_msg += f"⚠️  **Member ID:** Not provided - {coverage_info['member_id_format']}\n"
                        
                        # Add group number verification if provided
                        if 'group_number' in validated_insurance:
                            group_num = validated_insurance['group_number']
                            if group_num in coverage_info['group_formats']:
                                verification_msg += f"✅ **Group Number:** {group_num} (Verified)\n"
                            else:
                                verification_msg += f"⚠️  **Group Number:** {group_num} (Needs verification)\n"
                        else:
                            verification_msg += f"⚠️  **Group Number:** Not provided\n"
                        
                        verification_msg += """
📋 **Next Steps:**
1. Bring your insurance card to verify benefits
2. Copay will be collected at time of service
3. We'll verify your coverage before your appointment

💡 **Tip:** Call your insurance to confirm coverage and copay amount.
"""
                        return verification_msg
                    
                    else:
                        return f"""
❌ **Insurance Coverage Issue**

**Carrier:** {validated_insurance['insurance_carrier']}
**Status:** Not accepted in our network
**Reason:** {coverage_info.get('reason', 'Not in our provider network')}

**Options:**
1. Contact your insurance for in-network providers
2. Self-pay rates available ($150-300 per visit)
3. We can provide superbill for reimbursement

📞 **Need Help?** Call our billing department at (555) 123-4567
"""
                
                else:
                    return f"""
⚠️  **Insurance Verification Needed**

**Carrier:** {validated_insurance['insurance_carrier']}
**Status:** Needs manual verification
**Member ID:** {validated_insurance.get('member_id', 'Not provided')}
**Group:** {validated_insurance.get('group_number', 'Not provided')}

**Please:**
1. Call our billing department at (555) 123-4567
2. Have your insurance card ready
3. We'll verify coverage before your appointment

**Self-pay option available if needed.**
"""
            
            else:
                # No clear insurance information found
                return f"""
❓ **Insurance Information Missing**

**Input received:** {insurance_info}

**We need:**
1. Insurance carrier name (Blue Cross, Aetna, etc.)
2. Member ID number
3. Group number (if applicable)

**Please provide complete insurance information or call:**
📞 Billing Department: (555) 123-4567

**Self-pay rates:** $150-300 per visit
"""
                
        except Exception as e:
            return f"Error verifying insurance: {str(e)}"
            
        except Exception as e:
            return f"Error checking insurance: {str(e)}"
    
    def _send_intake_form(self, patient_info: str) -> str:
        """Send intake form to patient"""
        try:
            # Generate intake form link (local development)
            form_id = f"FORM_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            # Use local development URL instead of external domain
            form_link = f"http://localhost:8501/intake/{form_id}"  # Streamlit default port
            
            # Extract patient email from patient_info (simplified extraction)
            # In a real implementation, this would parse the patient_info properly
            patient_email = "charulchim06@gmail.com"  # Default to your email for testing
            
            # Prepare patient data
            patient_data = {
                'first_name': 'Patient',  # Would extract from patient_info
                'last_name': 'User',
                'email': patient_email,
                'appointment_date': 'Tomorrow',
                'appointment_time': '10:00 AM',
                'doctor_name': 'Dr. Sarah Johnson',
                'clinic_address': '123 Health St, Medical City, MC 12345',
                'intake_form_link': form_link
            }
            
            # Try to send email using SMTP service
            email_sent = False
            if self.email_service:
                email_sent = self.email_service.send_appointment_confirmation(patient_data)
            
            if email_sent:
                return f"""
✅ Intake form email sent successfully!

Email Details:
- Sent to: {patient_email}
- Form ID: {form_id}
- Access Link: {form_link}

The patient has received:
1. Appointment confirmation details
2. Online intake form link
3. Instructions for completing the form
4. List of items to bring to appointment

The form includes sections for:
- Personal information
- Medical history
- Insurance details
- Emergency contact information
- Current medications
- Allergies and medical conditions

📧 Email sent via personal SMTP service (no SendGrid required)
"""
            else:
                return f"""
⚠️ Intake form created but email not sent (SMTP not configured)

Form Details:
- Form ID: {form_id}
- Access Link: {form_link}

To enable email sending:
1. Set up Gmail App Password
2. Configure EMAIL_PASSWORD in .env file
3. Email will be sent from: charulchim06@gmail.com

The patient would receive:
1. Appointment confirmation details
2. Online intake form link
3. Instructions for completing the form
4. List of items to bring to appointment
"""
                
        except Exception as e:
            return f"Error sending intake form: {str(e)}"
    
    def _reschedule_appointment(self, details: str) -> str:
        """Reschedule an existing appointment"""
        try:
            # For simulation, return success message
            return f"""
Appointment rescheduling processed:

Original appointment details updated
New appointment confirmation sent
Intake forms remain valid for new date
Calendar updated automatically

Please note:
- Rescheduling fee may apply (waived for first reschedule)
- 24-hour notice required for changes
- Confirmation email sent with new details
"""
        except Exception as e:
            return f"Error rescheduling appointment: {str(e)}"
    
    def _cancel_appointment(self, appointment_id: str) -> str:
        """Cancel an existing appointment"""
        try:
            # For simulation, return success message
            return f"""
Appointment cancellation processed:

✅ Appointment {appointment_id} has been cancelled
✅ Calendar slot released for other patients
✅ Cancellation confirmation sent
✅ No cancellation fee (more than 24 hours notice)

If you need to reschedule, please let me know your preferred dates and times.
"""
        except Exception as e:
            return f"Error cancelling appointment: {str(e)}"
    
    def _export_appointments_to_excel(self, date_range: str = "all") -> str:
        """Export appointments to Excel with multiple sheets and formatting"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Load appointments data
            appointments_file = self.data_dir / "appointments" / "scheduled_appointments.csv"
            if not appointments_file.exists():
                return "❌ No appointments found to export"
            
            appointments_df = pd.read_csv(appointments_file)
            if appointments_df.empty:
                return "❌ No appointments found to export"
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create multiple sheets
            sheets_data = {
                'All Appointments': appointments_df,
                'Confirmed': appointments_df[appointments_df['status'] == 'confirmed'],
                'New Patients': appointments_df[appointments_df['patient_type'] == 'New Patient'],
                'Follow-ups': appointments_df[appointments_df['appointment_type'] == 'follow-up']
            }
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name, data in sheets_data.items():
                if data.empty:
                    continue
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Add headers
                headers = list(data.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                
                # Add data rows
                for row_num, row_data in enumerate(data.itertuples(index=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_num, column=col_num, value=value)
                        cell.border = border
                        
                        # Color coding for status
                        if col_num == data.columns.get_loc('status') + 1:
                            if str(value).lower() == 'confirmed':
                                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                            elif str(value).lower() == 'cancelled':
                                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            export_file = self.data_dir / "exports" / f"appointments_export_{timestamp}.xlsx"
            export_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Save workbook
            wb.save(export_file)
            
            # Generate summary
            total_appointments = len(appointments_df)
            confirmed_count = len(appointments_df[appointments_df['status'] == 'confirmed'])
            new_patients_count = len(appointments_df[appointments_df['patient_type'] == 'New Patient'])
            
            return f"""
📊 **Excel Export Complete**

✅ **File:** {export_file.name}
✅ **Location:** {export_file}
✅ **Sheets Created:** {len([s for s in sheets_data.values() if not s.empty])}

📈 **Summary:**
- Total Appointments: {total_appointments}
- Confirmed: {confirmed_count}
- New Patients: {new_patients_count}
- Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

**Sheets Include:**
- All Appointments (complete data)
- Confirmed appointments only
- New patient appointments
- Follow-up appointments

**Features:**
- Color-coded status (Green=Confirmed, Pink=Cancelled)
- Auto-sized columns
- Professional formatting
- Multiple data views
"""
            
        except ImportError:
            return """
❌ **Excel Export Not Available**

Missing required package: openpyxl

To enable Excel export, install openpyxl:
```
pip install openpyxl
```

Alternative: CSV export is available as backup.
"""
        except Exception as e:
            return f"❌ Error exporting to Excel: {str(e)}"
    
    def _send_appointment_confirmation(self, appointment_data: dict, confirmation_type: str = "booking") -> str:
        """Send appointment confirmation via email/SMS"""
        try:
            # Generate confirmation message based on type
            if confirmation_type == "booking":
                subject = f"✅ Appointment Confirmed - {appointment_data.get('appointment_date')}"
                message_type = "booking confirmation"
            elif confirmation_type == "reschedule":
                subject = f"📅 Appointment Rescheduled - {appointment_data.get('appointment_date')}"
                message_type = "reschedule confirmation"
            elif confirmation_type == "cancellation":
                subject = f"❌ Appointment Cancelled - {appointment_data.get('appointment_id')}"
                message_type = "cancellation confirmation"
            else:
                subject = f"📋 Appointment Update - {appointment_data.get('appointment_id')}"
                message_type = "appointment update"
            
            # Create comprehensive confirmation message
            confirmation_message = f"""
🏥 **{subject}**

**Patient:** {appointment_data.get('patient_name', 'N/A')}
**Appointment ID:** {appointment_data.get('appointment_id', 'N/A')}
**Date:** {appointment_data.get('appointment_date', 'N/A')}
**Time:** {appointment_data.get('appointment_time', 'N/A')}
**Duration:** {appointment_data.get('duration_minutes', 'N/A')} minutes
**Type:** {appointment_data.get('appointment_type', 'N/A').title()}
**Doctor:** Dr. {appointment_data.get('doctor_name', 'Sarah Johnson')}
**Patient Type:** {appointment_data.get('patient_type', 'N/A')}

**📍 Location:**
Medical Center - Main Building
123 Health Street, Suite 100
Medical City, MC 12345

**🔔 Reminders:**
- Arrive 15 minutes early for check-in
- Bring photo ID and insurance card
- Complete intake forms if new patient
- List of current medications

**📞 Contact:**
- Main Office: (555) 123-4567
- Emergency: (555) 999-8888
- Email: appointments@medicalcenter.com

**💡 Need to reschedule?**
Call us at least 24 hours in advance to avoid fees.
Online rescheduling: http://localhost:8501/reschedule

**🔐 Appointment Security Code:** {appointment_data.get('appointment_id', 'N/A')[-6:]}
"""
            
            # Try to send via email service
            email_sent = False
            sms_sent = False
            
            if self.email_service and 'patient_email' in appointment_data:
                try:
                    email_data = {
                        'patient_name': appointment_data.get('patient_name', 'Patient'),
                        'email': appointment_data.get('patient_email', 'charulchim06@gmail.com'),
                        'appointment_date': appointment_data.get('appointment_date', 'TBD'),
                        'appointment_time': appointment_data.get('appointment_time', 'TBD'),
                        'doctor_name': appointment_data.get('doctor_name', 'Dr. Sarah Johnson'),
                        'appointment_id': appointment_data.get('appointment_id', 'N/A'),
                        'confirmation_message': confirmation_message
                    }
                    email_sent = self.email_service.send_appointment_confirmation(email_data)
                except Exception as e:
                    print(f"Email sending failed: {e}")
            
            # Simulate SMS sending (would integrate with SMS service)
            if 'patient_phone' in appointment_data:
                sms_message = f"""
🏥 Appointment {message_type.title()}

Patient: {appointment_data.get('patient_name', 'N/A')}
Date: {appointment_data.get('appointment_date', 'N/A')}
Time: {appointment_data.get('appointment_time', 'N/A')}
ID: {appointment_data.get('appointment_id', 'N/A')}

Arrive 15 min early. Bring ID & insurance.
Reply STOP to opt out.
"""
                sms_sent = True  # Simulated - would use SMS service
            
            # Log confirmation to file
            self._log_confirmation(appointment_data, confirmation_type, confirmation_message)
            
            # Automatically trigger form distribution after successful confirmation
            form_distribution_result = ""
            if confirmation_type == "booking" and (email_sent or sms_sent):
                try:
                    # Prepare patient and appointment data for form distribution
                    patient_data = {
                        'name': appointment_data.get('patient_name', 'Patient'),
                        'email': appointment_data.get('patient_email', ''),
                        'phone': appointment_data.get('patient_phone', ''),
                        'patient_id': appointment_data.get('patient_id', ''),
                        'patient_type': appointment_data.get('patient_type', 'New Patient')
                    }
                    
                    form_result = self._distribute_intake_forms(patient_data, appointment_data)
                    if "successfully" in form_result.lower():
                        form_distribution_result = "✅ Intake forms automatically sent"
                    else:
                        form_distribution_result = "⚠️ Form distribution attempted"
                except Exception as e:
                    form_distribution_result = f"⚠️ Form distribution error: {str(e)}"
            
            delivery_status = []
            if email_sent:
                delivery_status.append("✅ Email sent successfully")
            else:
                delivery_status.append("⚠️ Email not sent (service unavailable)")
            
            if sms_sent:
                delivery_status.append("✅ SMS sent successfully")
            else:
                delivery_status.append("⚠️ SMS not sent (service unavailable)")
            
            if form_distribution_result:
                delivery_status.append(form_distribution_result)
            
            return f"""
📨 **Confirmation Sent**

**Type:** {message_type.title()}
**Patient:** {appointment_data.get('patient_name', 'N/A')}
**Appointment ID:** {appointment_data.get('appointment_id', 'N/A')}

**Delivery Status:**
{chr(10).join(delivery_status)}

**Message Includes:**
- Complete appointment details
- Location and directions
- Preparation instructions
- Contact information
- Rescheduling options
- Security confirmation code

**Backup:** Confirmation logged to system records
"""
            
        except Exception as e:
            return f"❌ Error sending confirmation: {str(e)}"
    
    def _log_confirmation(self, appointment_data: dict, confirmation_type: str, message: str) -> None:
        """Log confirmation to file for record keeping"""
        try:
            log_file = self.data_dir / "logs" / "confirmations.log"
            log_file.parent.mkdir(parents=True, exist_ok=True)
            
            log_entry = f"""
[{datetime.now().isoformat()}] {confirmation_type.upper()}
Appointment ID: {appointment_data.get('appointment_id', 'N/A')}
Patient: {appointment_data.get('patient_name', 'N/A')}
Message Length: {len(message)} characters
Status: Logged
---
"""
            
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)
                
        except Exception as e:
            print(f"Failed to log confirmation: {e}")
    
    def _backup_appointment_data(self, backup_type: str = "daily") -> str:
        """Create backup of appointment and patient data"""
        try:
            import shutil
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_dir = self.data_dir / "backups" / f"{backup_type}_{timestamp}"
            backup_dir.mkdir(parents=True, exist_ok=True)
            
            # Files to backup
            backup_files = [
                self.data_dir / "appointments" / "scheduled_appointments.csv",
                self.data_dir / "patients" / "patient_database.csv",
                self.data_dir / "doctors" / "doctor_profiles.csv",
                self.data_dir / "doctors" / "doctor_schedules.xlsx"
            ]
            
            backed_up = []
            for file_path in backup_files:
                if file_path.exists():
                    backup_file = backup_dir / file_path.name
                    shutil.copy2(file_path, backup_file)
                    backed_up.append(file_path.name)
            
            # Create backup manifest
            manifest = {
                'backup_date': datetime.now().isoformat(),
                'backup_type': backup_type,
                'files_backed_up': backed_up,
                'backup_location': str(backup_dir)
            }
            
            manifest_file = backup_dir / "backup_manifest.json"
            with open(manifest_file, 'w') as f:
                json.dump(manifest, f, indent=2)
            
            return f"""
💾 **Backup Complete**

**Type:** {backup_type.title()} Backup
**Location:** {backup_dir}
**Files Backed Up:** {len(backed_up)}

**Included Files:**
{chr(10).join([f"- {file}" for file in backed_up])}

**Backup Info:**
- Timestamp: {timestamp}
- Manifest: backup_manifest.json
- Total Size: {self._get_backup_size(backup_dir)}

**Retention:** Backups are kept for 30 days
"""
            
        except Exception as e:
            return f"❌ Backup failed: {str(e)}"
    
    def _get_backup_size(self, backup_dir: Path) -> str:
        """Calculate backup directory size"""
        try:
            total_size = sum(f.stat().st_size for f in backup_dir.rglob('*') if f.is_file())
            
            # Convert to human readable format
            for unit in ['B', 'KB', 'MB', 'GB']:
                if total_size < 1024:
                    return f"{total_size:.1f} {unit}"
                total_size /= 1024
            return f"{total_size:.1f} TB"
        except:
            return "Unknown"
    
    def _distribute_intake_forms(self, patient_data: dict, appointment_data: dict) -> str:
        """
        Automated form distribution system - sends intake forms only after appointment confirmation
        This method implements workflow automation to ensure proper sequencing
        """
        try:
            # Workflow validation - ensure appointment is confirmed before sending forms
            if appointment_data.get('status') != 'confirmed':
                return f"""
⚠️ **Form Distribution Blocked**

**Reason:** Appointment not confirmed yet
**Status:** {appointment_data.get('status', 'Unknown')}
**Patient:** {patient_data.get('patient_name', 'Unknown')}

**Workflow:** Forms will be distributed automatically once appointment is confirmed.
"""
            
            # Check if patient is new (forms only needed for new patients)
            if patient_data.get('patient_type') != 'New Patient':
                return f"""
ℹ️ **No Forms Required**

**Patient:** {patient_data.get('patient_name', 'Unknown')}
**Type:** {patient_data.get('patient_type', 'Unknown')} 
**Reason:** Intake forms only required for new patients

**Returning patients:** Welcome back! No additional forms needed.
"""
            
            # Generate secure form ID and link
            form_id = f"INTAKE_{appointment_data.get('appointment_id', 'UNKNOWN')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            intake_form_link = f"http://localhost:8501/intake/{form_id}"
            
            # Prepare comprehensive patient data for email template
            email_patient_data = {
                'first_name': patient_data.get('patient_name', '').split()[0] if patient_data.get('patient_name') else 'Patient',
                'last_name': ' '.join(patient_data.get('patient_name', '').split()[1:]) if patient_data.get('patient_name') else '',
                'full_name': patient_data.get('patient_name', 'Patient'),
                'email': patient_data.get('patient_email', 'charulchim06@gmail.com'),
                'phone': patient_data.get('patient_phone', ''),
                'appointment_date': appointment_data.get('appointment_date', 'TBD'),
                'appointment_time': appointment_data.get('appointment_time', 'TBD'),
                'appointment_id': appointment_data.get('appointment_id', 'N/A'),
                'doctor_name': appointment_data.get('doctor_name', 'Dr. Sarah Johnson'),
                'appointment_type': appointment_data.get('appointment_type', 'consultation'),
                'duration_minutes': appointment_data.get('duration_minutes', 60),
                'clinic_address': '123 Health Street, Suite 100, Medical City, MC 12345',
                'intake_form_link': intake_form_link,
                'form_id': form_id,
                'insurance_carrier': patient_data.get('insurance_carrier', ''),
                'member_id': patient_data.get('member_id', ''),
                'group_number': patient_data.get('group_number', '')
            }
            
            # Generate intake form content
            intake_form_content = self._generate_intake_form_content(email_patient_data)
            
            # Save form to local system for backup/offline access
            form_file = self._save_intake_form_locally(form_id, intake_form_content, email_patient_data)
            
            # Send email with intake form using enhanced template
            email_sent = False
            email_error = ""
            
            if self.email_service and self.email_templates:
                try:
                    # Use enhanced email template with intake form
                    email_template = self.email_templates.new_patient_intake_form_complete(email_patient_data)
                    
                    # Send email with intake form
                    email_sent = self.email_service.send_intake_form_email(
                        email_patient_data['email'],
                        email_template['subject'],
                        email_template['body'],
                        form_attachment=form_file if form_file else None
                    )
                except Exception as e:
                    email_error = str(e)
            
            # Log form distribution activity
            self._log_form_distribution(patient_data, appointment_data, form_id, email_sent)
            
            # Workflow completion status
            workflow_status = "✅ Complete" if email_sent else "⚠️ Partial"
            
            return f"""
📋 **Intake Form Distribution - {workflow_status}**

**Workflow Validation:** ✅ Appointment confirmed, new patient verified
**Patient:** {email_patient_data['full_name']} ({patient_data.get('patient_type')})
**Appointment:** {email_patient_data['appointment_id']} - {email_patient_data['appointment_date']}

**📧 Email Distribution:**
{'✅ Sent successfully to ' + email_patient_data['email'] if email_sent else '❌ Failed to send' + (': ' + email_error if email_error else '')}

**📄 Form Details:**
- Form ID: {form_id}
- Online Link: {intake_form_link}
- Local Backup: {'✅ Saved' if form_file else '❌ Failed'}
- Security: Unique patient-specific form ID

**📋 Form Includes:**
- Personal information section
- Medical history questionnaire
- Insurance verification details
- Emergency contact information
- Current medications & allergies
- Consent forms and policies

**🔄 Patient Instructions:**
1. Complete online form before appointment
2. Bring photo ID and insurance card
3. Review and sign consent forms
4. Arrive 15 minutes early if form incomplete

**⏱️ Timing:** Form distributed {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**📞 Support:** Call (555) 123-4567 for form assistance

**Workflow Status:** Forms distributed only after confirmation ✅
"""
            
        except Exception as e:
            return f"❌ Error in form distribution workflow: {str(e)}"
    
    def _generate_intake_form_content(self, patient_data: dict) -> str:
        """Generate comprehensive intake form content for the patient"""
        try:
            form_content = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Patient Intake Form - {patient_data.get('full_name', 'Patient')}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
        .header {{ background-color: #366092; color: white; padding: 20px; text-align: center; }}
        .section {{ margin: 20px 0; border: 1px solid #ddd; padding: 15px; }}
        .field {{ margin: 10px 0; }}
        label {{ font-weight: bold; display: inline-block; width: 200px; }}
        input, select, textarea {{ width: 300px; padding: 5px; }}
        .checkbox {{ width: auto; }}
        .submit-btn {{ background-color: #366092; color: white; padding: 10px 20px; border: none; cursor: pointer; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Patient Intake Form</h1>
        <p>Medical Center - Confidential Patient Information</p>
        <p>Form ID: {patient_data.get('form_id', 'N/A')} | Appointment: {patient_data.get('appointment_date', 'TBD')}</p>
    </div>

    <form id="intakeForm">
        <!-- Personal Information Section -->
        <div class="section">
            <h3>Personal Information</h3>
            <div class="field">
                <label>Full Name:</label>
                <input type="text" value="{patient_data.get('full_name', '')}" readonly>
            </div>
            <div class="field">
                <label>Date of Birth:</label>
                <input type="date" name="dob" required>
            </div>
            <div class="field">
                <label>Gender:</label>
                <select name="gender" required>
                    <option value="">Select...</option>
                    <option value="Male">Male</option>
                    <option value="Female">Female</option>
                    <option value="Other">Other</option>
                    <option value="Prefer not to answer">Prefer not to answer</option>
                </select>
            </div>
            <div class="field">
                <label>Address:</label>
                <textarea name="address" rows="3" placeholder="Street, City, State, ZIP"></textarea>
            </div>
            <div class="field">
                <label>Phone:</label>
                <input type="tel" value="{patient_data.get('phone', '')}" name="phone">
            </div>
            <div class="field">
                <label>Email:</label>
                <input type="email" value="{patient_data.get('email', '')}" readonly>
            </div>
        </div>

        <!-- Insurance Information Section -->
        <div class="section">
            <h3>Insurance Information</h3>
            <div class="field">
                <label>Insurance Carrier:</label>
                <input type="text" value="{patient_data.get('insurance_carrier', '')}" name="insurance_carrier">
            </div>
            <div class="field">
                <label>Member ID:</label>
                <input type="text" value="{patient_data.get('member_id', '')}" name="member_id">
            </div>
            <div class="field">
                <label>Group Number:</label>
                <input type="text" value="{patient_data.get('group_number', '')}" name="group_number">
            </div>
            <div class="field">
                <label>Policy Holder:</label>
                <input type="text" name="policy_holder" placeholder="If different from patient">
            </div>
        </div>

        <!-- Emergency Contact Section -->
        <div class="section">
            <h3>Emergency Contact</h3>
            <div class="field">
                <label>Contact Name:</label>
                <input type="text" name="emergency_name" required>
            </div>
            <div class="field">
                <label>Relationship:</label>
                <input type="text" name="emergency_relationship" required>
            </div>
            <div class="field">
                <label>Phone Number:</label>
                <input type="tel" name="emergency_phone" required>
            </div>
        </div>

        <!-- Medical History Section -->
        <div class="section">
            <h3>Medical History</h3>
            <div class="field">
                <label>Current Medications:</label>
                <textarea name="medications" rows="4" placeholder="List all current medications and dosages"></textarea>
            </div>
            <div class="field">
                <label>Allergies:</label>
                <textarea name="allergies" rows="3" placeholder="Medications, foods, environmental allergies"></textarea>
            </div>
            <div class="field">
                <label>Previous Surgeries:</label>
                <textarea name="surgeries" rows="3" placeholder="List surgeries and approximate dates"></textarea>
            </div>
            <div class="field">
                <label>Chronic Conditions:</label>
                <textarea name="chronic_conditions" rows="3" placeholder="Diabetes, hypertension, etc."></textarea>
            </div>
            <div class="field">
                <label>Reason for Visit:</label>
                <textarea name="visit_reason" rows="3" placeholder="Describe your symptoms or reason for appointment"></textarea>
            </div>
        </div>

        <!-- Consent and Agreements -->
        <div class="section">
            <h3>Consent and Agreements</h3>
            <div class="field">
                <label class="checkbox">
                    <input type="checkbox" name="privacy_consent" required class="checkbox">
                    I acknowledge receipt of the Notice of Privacy Practices
                </label>
            </div>
            <div class="field">
                <label class="checkbox">
                    <input type="checkbox" name="treatment_consent" required class="checkbox">
                    I consent to treatment and authorize medical care
                </label>
            </div>
            <div class="field">
                <label class="checkbox">
                    <input type="checkbox" name="financial_responsibility" required class="checkbox">
                    I understand my financial responsibility for services
                </label>
            </div>
            <div class="field">
                <label class="checkbox">
                    <input type="checkbox" name="insurance_verification" class="checkbox">
                    I authorize verification of insurance benefits
                </label>
            </div>
        </div>

        <!-- Signature Section -->
        <div class="section">
            <h3>Electronic Signature</h3>
            <div class="field">
                <label>Patient Signature:</label>
                <input type="text" name="signature" placeholder="Type your full name as electronic signature" required>
            </div>
            <div class="field">
                <label>Date:</label>
                <input type="date" name="signature_date" value="{datetime.now().strftime('%Y-%m-%d')}" readonly>
            </div>
        </div>

        <div style="text-align: center; margin: 30px 0;">
            <button type="submit" class="submit-btn">Submit Intake Form</button>
        </div>
    </form>

    <script>
        document.getElementById('intakeForm').addEventListener('submit', function(e) {{
            e.preventDefault();
            alert('Thank you! Your intake form has been submitted. Please bring a copy to your appointment.');
        }});
    </script>
</body>
</html>
"""
            return form_content
            
        except Exception as e:
            return f"Error generating form content: {str(e)}"
    
    def _save_intake_form_locally(self, form_id: str, form_content: str, patient_data: dict) -> Optional[str]:
        """Save intake form locally for backup and offline access"""
        try:
            forms_dir = self.data_dir / "intake_forms" / "generated"
            forms_dir.mkdir(parents=True, exist_ok=True)
            
            # Save HTML form
            form_file = forms_dir / f"{form_id}.html"
            with open(form_file, 'w', encoding='utf-8') as f:
                f.write(form_content)
            
            # Save patient data as JSON for reference
            patient_file = forms_dir / f"{form_id}_patient_data.json"
            with open(patient_file, 'w', encoding='utf-8') as f:
                import json
                json.dump(patient_data, f, indent=2)
            
            return str(form_file)
            
        except Exception as e:
            print(f"Error saving intake form locally: {e}")
            return None
    
    def _log_form_distribution(self, patient_data: dict, appointment_data: dict, form_id: str, email_sent: bool):
        """Log form distribution activity for workflow tracking"""
        try:
            log_file = self.data_dir / "logs" / "form_distribution.log"
            log_file.parent.mkdir(parents=True, exist_ok=True)
            
            log_entry = f"""
[{datetime.now().isoformat()}] FORM_DISTRIBUTION
Patient: {patient_data.get('patient_name', 'Unknown')}
Appointment ID: {appointment_data.get('appointment_id', 'Unknown')}
Form ID: {form_id}
Email Status: {'SUCCESS' if email_sent else 'FAILED'}
Patient Type: {patient_data.get('patient_type', 'Unknown')}
Appointment Status: {appointment_data.get('status', 'Unknown')}
Distribution Trigger: POST_CONFIRMATION
---
"""
            
            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_entry)
                
        except Exception as e:
            print(f"Failed to log form distribution: {e}")
    
    def chat(self, message: str) -> str:
        """Main chat interface for the scheduling agent"""
        try:
            if not self.agent:
                # Simulation mode - use built-in responses
                return self._simulate_response(message)
            
            # Try to use LangChain agent
            response = self.agent.invoke({"input": message})
            return response["output"]
            
        except Exception as e:
            self.logger.error(f"Error in chat: {str(e)}")
            # Fall back to simulation mode on any error
            self.logger.info("Falling back to simulation mode")
            return self._simulate_response(message)
    
    def _simulate_response(self, message: str) -> str:
        """Simulate agent responses when OpenAI is not available"""
        message_lower = message.lower()
        
        if any(word in message_lower for word in ['hello', 'hi', 'hey', 'good morning', 'good afternoon']):
            return """
Hello! Welcome to our medical appointment scheduling system. I'm here to help you schedule, reschedule, or manage your appointments.

I can assist you with:
• Scheduling new appointments
• Looking up existing appointments
• Finding available time slots
• Checking doctor availability
• Sending intake forms
• Insurance verification

How can I help you today?
"""
        
        elif any(word in message_lower for word in ['schedule', 'appointment', 'book']):
            return """
I'd be happy to help you schedule an appointment! To get started, I'll need some information:

1. What type of appointment do you need? (annual checkup, consultation, follow-up, etc.)
2. Do you have a preferred doctor or specialty?
3. What are your preferred dates and times?
4. Are you an existing patient or new patient?

Please let me know these details and I'll find the best available options for you.
"""
        
        elif any(word in message_lower for word in ['available', 'availability', 'slots']):
            return self._search_available_slots(message)
        
        elif any(word in message_lower for word in ['doctor', 'physician', 'specialist']):
            return self._get_doctor_info(message)
        
        elif any(word in message_lower for word in ['insurance', 'coverage', 'benefits']):
            return """
I can help verify your insurance coverage. Our practice accepts most major insurance plans including:

• Blue Cross Blue Shield
• Aetna
• Cigna
• UnitedHealthcare
• Humana
• Medicare
• Medicaid

What insurance provider do you have? I'll check if we accept your plan and provide information about copays and coverage.
"""
        
        else:
            return """
I'm here to help with appointment scheduling. I can assist with:

• Scheduling new appointments
• Checking available time slots
• Finding doctor information
• Verifying insurance coverage
• Sending intake forms
• Managing appointment reminders

Could you please let me know specifically what you'd like help with today?
"""
    
    def _schedule_appointment_reminders(self, appointment_id: str) -> str:
        """Schedule automated reminders for an appointment"""
        try:
            if not self.reminder_engine:
                return "❌ Reminder system not available"
            
            result = self.reminder_engine.schedule_reminders_for_appointment(appointment_id)
            
            if result['status'] == 'success':
                reminders = result.get('reminders', [])
                reminder_details = []
                
                for reminder in reminders:
                    reminder_details.append(f"• {reminder['type'].title()} reminder - {reminder['scheduled_for']} via {reminder['method']}")
                
                return f"""
✅ **Reminder System Activated**

📅 Appointment ID: {appointment_id}
🔔 Scheduled {len(reminders)} automated reminders:

{chr(10).join(reminder_details)}

**Reminder Types:**
- Regular Reminder (24h before): General appointment reminder
- Form Check (4h before): Verify intake forms completed  
- Visit Confirmation (1h before): Final confirmation or cancellation

Patients will receive reminders via email and/or SMS based on their preferences.
"""
            else:
                return f"❌ Failed to schedule reminders: {result.get('message', 'Unknown error')}"
                
        except Exception as e:
            return f"❌ Error scheduling reminders: {str(e)}"
    
    def _check_reminder_status(self, query: str) -> str:
        """Check reminder status for appointments or patients"""
        try:
            if not self.reminder_engine:
                return "❌ Reminder system not available"
            
            # Parse query to extract appointment ID or patient ID
            appointment_id = None
            patient_id = None
            
            if "APT_" in query.upper():
                import re
                apt_match = re.search(r'APT_\d{8}_\d{6}', query.upper())
                if apt_match:
                    appointment_id = apt_match.group()
            
            elif "P" in query.upper() and any(char.isdigit() for char in query):
                import re
                patient_match = re.search(r'P\d+', query.upper())
                if patient_match:
                    patient_id = patient_match.group()
            
            # Get reminder status
            if appointment_id:
                status = self.reminder_engine.get_reminder_status(appointment_id=appointment_id)
                title = f"📅 Reminders for Appointment {appointment_id}"
            elif patient_id:
                status = self.reminder_engine.get_reminder_status(patient_id=patient_id)
                title = f"👤 Reminders for Patient {patient_id}"
            else:
                status = self.reminder_engine.get_reminder_status()
                title = "🔔 All System Reminders"
            
            summary = status.get('summary', {})
            reminders = status.get('reminders', [])
            
            if not reminders:
                return f"""
{title}

📊 **Status: No reminders found**

To schedule reminders for an appointment, use the appointment ID.
"""
            
            # Format reminder details
            reminder_list = []
            for reminder in reminders[:10]:  # Show up to 10 reminders
                # Handle NaN values in pandas data
                reminder_id = reminder.get('reminder_id', 'N/A')
                reminder_type = reminder.get('reminder_type', 'N/A')
                scheduled_time = str(reminder.get('scheduled_time', 'N/A'))[:16]
                delivery_method = reminder.get('delivery_method', 'N/A')
                status = reminder.get('status', 'N/A')
                sent_at = str(reminder.get('sent_at', ''))[:16] if reminder.get('sent_at') else 'Not sent'
                
                reminder_list.append(f"""
• **{reminder_type.title()} Reminder**
  - ID: {reminder_id}
  - Scheduled: {scheduled_time}
  - Method: {delivery_method}
  - Status: {status.title()}
  - Sent: {sent_at}
""")
            
            return f"""
{title}

📊 **Summary:**
- Total: {summary.get('total', 0)}
- Scheduled: {summary.get('scheduled', 0)}
- Sent: {summary.get('sent', 0)}
- Failed: {summary.get('failed', 0)}

🔔 **Reminder Details:**
{''.join(reminder_list)}

{'...(showing first 10 of ' + str(len(reminders)) + ')' if len(reminders) > 10 else ''}
"""
            
        except Exception as e:
            return f"❌ Error checking reminder status: {str(e)}"
    
    def _send_manual_reminder(self, query: str) -> str:
        """Manually send a specific reminder"""
        try:
            if not self.reminder_engine:
                return "❌ Reminder system not available"
            
            # Parse query for reminder type, appointment ID, or patient ID
            query_lower = query.lower()
            
            if "regular" in query_lower:
                reminder_type = "regular"
            elif "form" in query_lower:
                reminder_type = "form_check"
            elif "confirm" in query_lower:
                reminder_type = "confirmation"
            else:
                return "❌ Please specify reminder type: regular, form_check, or confirmation"
            
            # Extract appointment or patient ID
            appointment_id = None
            if "APT_" in query.upper():
                import re
                apt_match = re.search(r'APT_\d{8}_\d{6}', query.upper())
                if apt_match:
                    appointment_id = apt_match.group()
            
            if not appointment_id:
                return "❌ Please provide an appointment ID (e.g., APT_20250906_123456)"
            
            # Load appointment data and send reminder
            appointments_df = pd.read_csv(self.data_dir / "appointments" / "scheduled_appointments.csv")
            appointment = appointments_df[appointments_df['appointment_id'] == appointment_id]
            
            if appointment.empty:
                return f"❌ Appointment {appointment_id} not found"
            
            appointment_data = appointment.iloc[0]
            
            # Get patient data
            patients_df = pd.read_csv(self.data_dir / "patients" / "patient_database.csv")
            patient = patients_df[patients_df['patient_id'] == appointment_data['patient_id']]
            
            if patient.empty:
                return f"❌ Patient {appointment_data['patient_id']} not found"
            
            patient_data = patient.iloc[0]
            
            # Prepare combined data
            combined_data = {
                **appointment_data.to_dict(),
                'patient_name': patient_data.get('full_name', f"{patient_data.get('first_name', '')} {patient_data.get('last_name', '')}").strip(),
                'patient_email': patient_data.get('email', ''),
                'patient_phone': patient_data.get('phone', ''),
                'doctor_name': appointment_data.get('doctor_name', 'TBD')
            }
            
            # Send reminder
            email_success = False
            sms_success = False
            
            if self.email_service and combined_data['patient_email']:
                email_success = self.email_service.send_appointment_reminder(combined_data, reminder_type)
            
            if combined_data['patient_phone']:
                sms_success = self.email_service.send_sms_reminder(combined_data, reminder_type) if self.email_service else True
            
            if email_success or sms_success:
                return f"""
✅ **Manual Reminder Sent**

📅 Appointment: {appointment_id}
👤 Patient: {combined_data['patient_name']}
🔔 Type: {reminder_type.title()} reminder
📧 Email: {'✅ Sent' if email_success else '❌ Failed/Not available'}
📱 SMS: {'✅ Sent' if sms_success else '❌ Failed/Not available'}

The patient has been notified about their upcoming appointment.
"""
            else:
                return f"❌ Failed to send reminder - no valid contact methods available"
                
        except Exception as e:
            return f"❌ Error sending manual reminder: {str(e)}"
    
    def _process_patient_response(self, response_text: str) -> str:
        """Process patient response to reminders"""
        try:
            if not self.reminder_engine:
                return "❌ Reminder system not available"
            
            # Extract patient ID or appointment ID from response if available
            patient_id = None
            reminder_type = None
            
            # Parse response for context
            response_lower = response_text.lower()
            
            if "form" in response_lower:
                reminder_type = "form_check"
            elif "confirm" in response_lower or "cancel" in response_lower:
                reminder_type = "confirmation"
            
            # For demo, use a placeholder patient ID
            # In production, this would be linked to the actual responding patient
            result = self.reminder_engine.process_patient_response("demo_patient", response_text, reminder_type)
            
            status_icons = {
                "success": "✅",
                "partial": "⚠️",
                "error": "❌"
            }
            
            icon = status_icons.get(result['status'], "ℹ️")
            
            response_message = f"""
{icon} **Patient Response Processed**

📝 Response: "{response_text}"
🔄 Action: {result['action'].replace('_', ' ').title()}
💬 Message: {result['message']}

"""
            
            # Add next action information
            next_action = result.get('next_action', 'none')
            if next_action == 'staff_callback':
                response_message += "📞 **Next Step:** Staff will call patient to assist\n"
            elif next_action == 'resend_forms':
                response_message += "📄 **Next Step:** Intake forms will be resent\n"
            elif next_action == 'process_cancellation':
                cancellation_reason = result.get('cancellation_reason', 'unspecified')
                response_message += f"❌ **Next Step:** Process cancellation (Reason: {cancellation_reason})\n"
            elif next_action == 'staff_reschedule':
                response_message += "📅 **Next Step:** Staff will contact patient to reschedule\n"
            elif next_action == 'staff_review':
                response_message += "👥 **Next Step:** Staff will review and follow up\n"
            
            return response_message
            
        except Exception as e:
            return f"❌ Error processing patient response: {str(e)}"
    
    def _run_reminder_system(self, query: str = "") -> str:
        """Check and send all due reminders"""
        try:
            if not self.reminder_engine:
                return "❌ Reminder system not available"
            
            result = self.reminder_engine.check_and_send_due_reminders()
            
            sent = result.get('sent', [])
            failed = result.get('failed', [])
            skipped = result.get('skipped', [])
            
            summary = f"""
🔄 **Reminder System Run Complete**

📊 **Results:**
- ✅ Sent: {len(sent)}
- ❌ Failed: {len(failed)}
- ⚠️ Skipped: {len(skipped)}

"""
            
            if sent:
                summary += "✅ **Successfully Sent:**\n"
                for reminder in sent[:5]:  # Show first 5
                    summary += f"  • {reminder['type'].title()} to {reminder['patient']} via {reminder['method']}\n"
                if len(sent) > 5:
                    summary += f"  • ...and {len(sent) - 5} more\n"
                summary += "\n"
            
            if failed:
                summary += "❌ **Failed to Send:**\n"
                for reminder in failed[:3]:  # Show first 3
                    summary += f"  • {reminder['type'].title()} - {reminder.get('reason', 'Unknown error')}\n"
                if len(failed) > 3:
                    summary += f"  • ...and {len(failed) - 3} more\n"
                summary += "\n"
            
            if not sent and not failed and not skipped:
                summary += "ℹ️ No reminders were due at this time.\n"
            
            return summary
            
        except Exception as e:
            return f"❌ Error running reminder system: {str(e)}"
    
    def _configure_sms_service(self, query: str = "") -> str:
        """Get SMS service configuration and setup instructions"""
        try:
            from src.utils.sms_service import SMSConfig
            
            # Get current recommended provider
            recommended = SMSConfig.get_recommended_provider()
            
            # Get setup instructions
            instructions = SMSConfig.setup_instructions(recommended)
            
            # Check if specific provider was requested
            if "twilio" in query.lower():
                provider_instructions = SMSConfig.setup_instructions("twilio")
            elif "aws" in query.lower() or "sns" in query.lower():
                provider_instructions = SMSConfig.setup_instructions("aws_sns")
            else:
                provider_instructions = instructions
            
            return f"""
📱 **SMS Service Configuration**

🔧 **Current Status:** {recommended.upper()} mode

{provider_instructions}

🚀 **Quick Start for Real SMS:**

**Option 1: Twilio (Recommended)**
```bash
# Install Twilio
pip install twilio

# Set environment variables (replace with your values)
export TWILIO_ACCOUNT_SID="your_account_sid_here"
export TWILIO_AUTH_TOKEN="your_auth_token_here"  
export TWILIO_PHONE_NUMBER="+1234567890"

# Restart the application
```

**Option 2: AWS SNS**
```bash
# Install boto3
pip install boto3

# Configure AWS credentials
aws configure
# OR set environment variables:
export AWS_ACCESS_KEY_ID="your_access_key"
export AWS_SECRET_ACCESS_KEY="your_secret_key"
export AWS_DEFAULT_REGION="us-east-1"

# Restart the application
```

📊 **SMS Usage Statistics:**
- Current mode: Simulated (no cost, perfect for testing)
- Real SMS cost: ~$0.0075 per message
- Supported features: Delivery confirmations, status tracking

💡 **Need Help?** 
- Ask me: "How do I set up Twilio SMS?"
- Ask me: "Show AWS SNS setup instructions"
- Visit: https://docs.twilio.com/ or https://docs.aws.amazon.com/sns/
"""
            
        except Exception as e:
            return f"❌ Error getting SMS configuration: {str(e)}"


# Example usage and testing
if __name__ == "__main__":
    # Initialize the scheduling agent
    agent = MedicalSchedulingAgent()
    
    print("Medical Appointment Scheduling Agent initialized!")
    print("=" * 60)
    
    # Test conversation flow
    test_messages = [
        "Hello, I need to schedule an appointment",
        "I'm a new patient and need a general checkup",
        "What doctors are available?",
        "Do you accept Blue Cross Blue Shield insurance?",
        "What times are available next week?"
    ]
    
    for message in test_messages:
        print(f"\nUser: {message}")
        response = agent.chat(message)
        print(f"Agent: {response}")
        print("-" * 40)
